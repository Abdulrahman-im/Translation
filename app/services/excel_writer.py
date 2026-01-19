"""Excel file generation service."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Tuple


def create_excel_file(
    translations: List[Tuple[int, str, str]],
    output_path: str
) -> str:
    """
    Create an Excel file with translations.

    Args:
        translations: List of (slide_number, original_text, translated_text) tuples
        output_path: Path where the Excel file will be saved

    Returns:
        Path to the created Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"

    # Define headers
    headers = ["Slide Number", "Original Phrase", "Translation"]

    # Style for headers
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for row_num, (slide_num, original, translation) in enumerate(translations, start=2):
        ws.cell(row=row_num, column=1, value=slide_num)
        ws.cell(row=row_num, column=2, value=original)
        ws.cell(row=row_num, column=3, value=translation)

    # Auto-fit columns
    column_widths = [15, 50, 50]  # Default widths
    for col_num, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Set right-to-left alignment for Arabic translation column
    for row in range(2, len(translations) + 2):
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")

    # Save workbook
    wb.save(output_path)

    return output_path
